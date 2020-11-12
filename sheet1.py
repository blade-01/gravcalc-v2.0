#this file appends titles to the spreadsheet

from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, Font, Color, Border, colors, Alignment, Side, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
import pandas as pd
from pandas import DataFrame 
from openpyxl.drawing.image import Image

#activate a spreadsheet
gravcalc = Workbook()
sheet_one = gravcalc.active
sheet_one.title = 'Workings'

#create an extra sheet for the formulae
sheet_two = gravcalc.create_sheet('Formulae',1)


#format the spreadsheet
bold_font = Font(bold=True, color=colors.BLUE, size=10)
font_alignment = Alignment(horizontal='center', vertical='center')
font_border = Border(bottom=Side(border_style='thin'))
sheet_one.freeze_panes = "B1"
#add titles to the spreadsheet
sheet_one["A1"] = "GRAVITY STATION"
sheet_one["A1"].font = bold_font
sheet_one["A3"] = "FIRST BASE STATION"
sheet_one["A3"].font = bold_font
sheet_one["A4"] = "LAST BASE STATION"
sheet_one["A4"].font = bold_font

sheet_one["B1"] = "TIME(hr)"
sheet_one["B1"].font = bold_font
sheet_one["B2"] = "(hr)" 
sheet_one["B2"].font = bold_font

sheet_one["C1"] = "TIME(min)"
sheet_one["C1"].font = bold_font
sheet_one["C2"] = "(min)" 
sheet_one["C2"].font = bold_font

sheet_one["D1"] = "ELEVATION(cm)"
sheet_one["D1"].font = bold_font
sheet_one["D2"] = "(cm)"
sheet_one["D2"].font = bold_font

sheet_one["E1"] = "ELEVATION(m)"
sheet_one["E1"].font = bold_font
sheet_one["E2"] = "(m)"
sheet_one["E2"].font = bold_font

sheet_one["F1"] = "LATITUDE(degrees)"
sheet_one["F1"].font = bold_font
sheet_one["F2"] = "Degrees"
sheet_one["F2"].font = bold_font

sheet_one["G1"] = "LATITUDE(min)"
sheet_one["G1"].font = bold_font
sheet_one["G2"] = "Minutes"
sheet_one["G2"].font = bold_font

sheet_one["H1"] = "LATITUDE(sec)"
sheet_one["H1"].font = bold_font
sheet_one["H2"] = "Seconds"
sheet_one["H2"].font = bold_font

sheet_one["I1"] = "LONGITUDE(degrees)"
sheet_one["I1"].font = bold_font
sheet_one["I2"] = "Degrees"
sheet_one["I2"].font = bold_font

sheet_one["J1"] = "LONGITUDE(min)"
sheet_one["J1"].font = bold_font
sheet_one["J2"] = "Minutes"
sheet_one["J2"].font = bold_font

sheet_one["K1"] = "LONGITUDE(sec)"
sheet_one["K1"].font = bold_font
sheet_one["K2"] = "Seconds"
sheet_one["K2"].font = bold_font

sheet_one["L1"] = "OBSERVED GRAVITY"
sheet_one["L1"].font = bold_font
sheet_one["L2"] = "(dial)"
sheet_one["L2"].font = bold_font

sheet_one["M1"] = "OBSERVED GRAVITY(mgal)"
sheet_one["M1"].font = bold_font
sheet_one["M2"] = "(mGal)"
sheet_one["M2"].font = bold_font

sheet_one["N1"] = "CHANGE IN OBSERVED GRAVITY(mgal)"
sheet_one["N1"].font = bold_font
sheet_one["N2"] = "(mGal)"
sheet_one["N2"].font = bold_font

sheet_one["O1"] = "TOTAL TIME"
sheet_one["O1"].font = bold_font
sheet_one["O2"] = "(min)"
sheet_one["O2"].font = bold_font

sheet_one["P1"] = "CHANGE IN TIME(min)"
sheet_one["P1"].font = bold_font
sheet_one["P2"] = "(min)" 
sheet_one["P2"].font = bold_font

sheet_one["Q1"] = "CHANGE IN ELEVATION(m)"
sheet_one["Q1"].font = bold_font
sheet_one["Q2"] = "(m)" 
sheet_one["Q2"].font = bold_font

sheet_one["R1"] = "LATITUDE(DD)"
sheet_one["R1"].font = bold_font
sheet_one["R2"] = "(DECIMAL DEGREES)"
sheet_one["R2"].font = bold_font

sheet_one["S1"] = "LONGITUDE(DD)"
sheet_one["S1"].font = bold_font
sheet_one["S2"] = "(DECIMAL DEGREES)"
sheet_one["S2"].font = bold_font

sheet_one["T1"] = "LATITUDE CORRECTION"
sheet_one["T1"].font = bold_font

sheet_one["U1"] = "DRIFT CORRECTION"
sheet_one["U1"].font = bold_font

sheet_one["V1"] = "DRIFT RATE"
sheet_one["V1"].font = bold_font

sheet_one["W1"] = "CORRECTED GRAVITY"
sheet_one["W1"].font = bold_font
sheet_one["W2"] = "(mGal)"
sheet_one["W2"].font = bold_font

sheet_one["X1"] = "FREE AIR CORRECTION"
sheet_one["X1"].font = bold_font
sheet_one["X2"] = "(mGal)"
sheet_one["X2"].font = bold_font

sheet_one["Y1"] = "FREE AIR ANOMALY"
sheet_one["Y1"].font = bold_font
sheet_one["Y2"] = "(mGal)"
sheet_one["Y2"].font = bold_font

sheet_one["Z1"] = "BOUGUER CORRECTION"
sheet_one["Z1"].font = bold_font
sheet_one["Z2"] = "(mGal)"
sheet_one["Z2"].font = bold_font

sheet_one["AA1"] = "BOUGUER ANOMALY"
sheet_one["AA1"].font = bold_font
sheet_one["AA2"] = "(mGal)"
sheet_one["AA2"].font = bold_font

#appending a picture to the second sheet
#create header
formuale_pic = Image("formulae.png")
formuale_pic.height = 400
formuale_pic.width = 550

sheet_two.add_image(formuale_pic, "H5")
#save the sheet
gravcalc.save(filename= "GravCalc.xlsx")
