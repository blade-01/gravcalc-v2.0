from openpyxl import Workbook, load_workbook
from pandas import DataFrame 
import pandas as pd

#activate a spreadsheet
gravcalc = Workbook()
sheet_one = gravcalc.active
gravcalc.save(filename= "GravCalc.xlsx")

#append values
edited_file = load_workbook(filename="GravCalc.xlsx")
data = edited_file.active
data.auto_filter.ref = "A1:O100"

#append the cell titles to the loaded spread sheet
data["A1"] = "GRAVITY STATION"
data["B1"] = "DATE (mm/dd/yyyy)"
data["C1"] = "TIME (HOURS)"
data["D1"] = "TIME(MINUTES)" 
data["E1"] = "DURATION(h)"
data["F1"] = "LATITUDE (degrees)"
data["F2"] = "Degrees"
data["G1"] = "LATITUDE(minutes)"
data["G2"] = "Minutes"
data["H1"] = "LATITUDE(seconds)"
data["H2"] = "Seconds"
data["I1"] = "LONGITUDE(degrees)"
data["I2"] = "Degrees"
data["J1"] = "LONGITUDE(minutes)"
data["J2"] = "Minutes"
data["K1"] = "LONGITUDE(seconds)"
data["K2"] = "Seconds"
data["L1"] = "LATITUDE (DD)"
data["M1"] = "LONGITUDE (DD)"
data["N1"] = "ELLIPSOID HEIGHT (m)"
data["O1"] = "LACOSTE-ROMBERG METER"
data["P1"] = "COUNTER AND CALIBRATED (mGal)"
data["Q1"] = "LACOSTE-ROMBERG, MEASURED (mGal)"
data["R1"] = "WORDEN METER, COUNTER (DIAL READING)"
data["S1"] = "CALIBRATED (mGal)"
data["T1"] = "SCINTREX METER, DRIFT AND TIDE CORRECTED (mGal)"
data["U1"] = "RAW OBSERVED GRAVITY (mGal)"
data["V1"] = "TIDE (mGal)"
data["W1"] = "TIDE CORRECTED (mGal)"
data["X1"] = "METER DRIFT (mGal)"
data["Y1"] = "DRIFT AND TIDE CORRECTED (mGal)"
data["Z1"] = "DC SHITF (mGal)"
data["AA1"] = "THEORETICAL GRAVITY (mGal)"
data["AB1"] = "HEIGHT CORRECTED (mGal)"
data["AC1"] = "ATM EFFECT (mGal)"
data["AD1"] = "BOUGUER SPHERICAL CAP (mGal)"
data["AE1"] = "TERRAIN CORRECTION (mGal)"
data["AF1"] = "OBSERVED ABSOLUTE GRAVITY (mGal)"
data["AG1"] = "COMPLETE BOUGUER ANOMALY (mGal)"

#FORMULAE
#save the edited spreadsheet
edited_file.save(filename="GravCalc.xlsx")